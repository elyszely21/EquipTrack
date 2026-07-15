from datetime import timedelta
import json
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Count, Sum, F
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.utils import timezone

from accounts.models import UserProfile
from .models import Equipment, BorrowRequest, BorrowRequestItem, ReturnRecord, AuditLog
from .forms import (
    EquipmentForm,
    BorrowRequestItemFormSet,
    ReturnRecordForm,
    EditProfileForm,
)
from .utils import log_action


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_admin_or_staff(user):
    return (
        user.is_authenticated
        and hasattr(user, "profile")
        and user.profile.role in (UserProfile.ROLE_ADMIN, UserProfile.ROLE_STAFF)
        and user.profile.status == UserProfile.STATUS_ACTIVE
    )


def _is_admin(user):
    return (
        user.is_authenticated
        and hasattr(user, "profile")
        and user.profile.role == UserProfile.ROLE_ADMIN
        and user.profile.status == UserProfile.STATUS_ACTIVE
    )


def _paginate(request, queryset, per_page=10):
    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get("page")
    return paginator.get_page(page_number)


def _monthly_counts(queryset, date_field, months=6):
    """Return labels and counts for the last `months` calendar months."""
    today = timezone.localdate()

    rows = (
        queryset.filter(**{f"{date_field}__date__gte": today.replace(day=1) - timedelta(days=months * 31)})
        .annotate(month=TruncMonth(date_field))
        .values("month")
        .annotate(total=Count("pk"))
        .order_by("month")
    )
    lookup = {
        row["month"].strftime("%Y-%m"): row["total"]
        for row in rows
        if row["month"] is not None
    }

    labels = []
    data = []
    cursor = today.replace(day=1)
    month_list = []
    for _ in range(months):
        month_list.append(cursor)
        if cursor.month == 1:
            cursor = cursor.replace(year=cursor.year - 1, month=12)
        else:
            cursor = cursor.replace(month=cursor.month - 1)
    month_list.reverse()

    for month_date in month_list:
        key = month_date.strftime("%Y-%m")
        labels.append(month_date.strftime("%b %Y"))
        data.append(lookup.get(key, 0))

    return labels, data


def _monthly_return_counts(months=6):
    today = timezone.localdate()
    start = today.replace(day=1) - timedelta(days=months * 31)

    rows = (
        ReturnRecord.objects.filter(return_date__gte=start, return_date__isnull=False)
        .annotate(month=TruncMonth("return_date"))
        .values("month")
        .annotate(total=Count("pk"))
        .order_by("month")
    )
    lookup = {
        row["month"].strftime("%Y-%m"): row["total"]
        for row in rows
        if row["month"] is not None
    }

    labels = []
    data = []
    cursor = today.replace(day=1)
    month_list = []
    for _ in range(months):
        month_list.append(cursor)
        if cursor.month == 1:
            cursor = cursor.replace(year=cursor.year - 1, month=12)
        else:
            cursor = cursor.replace(month=cursor.month - 1)
    month_list.reverse()

    for month_date in month_list:
        key = month_date.strftime("%Y-%m")
        labels.append(month_date.strftime("%b %Y"))
        data.append(lookup.get(key, 0))

    return labels, data


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@login_required
def dashboard(request):
    profile = request.user.profile

    # Use single aggregated query for equipment stats
    equipment_stats = (
        Equipment.objects
        .values('status')
        .annotate(count=Count('pk'))
    )
    stats_dict = {item['status']: item['count'] for item in equipment_stats}
    
    total_equipment = Equipment.objects.count()
    available_equipment = stats_dict.get(Equipment.STATUS_AVAILABLE, 0)
    unavailable_equipment = stats_dict.get(Equipment.STATUS_UNAVAILABLE, 0)
    maintenance_equipment = stats_dict.get(Equipment.STATUS_MAINTENANCE, 0)
    
    # Inventory alerts - use single query
    low_stock_equipment = Equipment.objects.filter(
        quantity_available__lte=5,
        quantity_available__gt=0
    ).count()
    out_of_stock_equipment = Equipment.objects.filter(
        quantity_available=0
    ).count()

    # Use single aggregated query for borrow stats
    borrow_stats = (
        BorrowRequest.objects
        .values('status')
        .annotate(count=Count('pk'))
    )
    borrow_stats_dict = {item['status']: item['count'] for item in borrow_stats}
    
    pending_borrows = borrow_stats_dict.get(BorrowRequest.STATUS_PENDING, 0)
    approved_borrows = borrow_stats_dict.get(BorrowRequest.STATUS_APPROVED, 0)
    rejected_borrows = borrow_stats_dict.get(BorrowRequest.STATUS_REJECTED, 0)
    returned_borrows = borrow_stats_dict.get(BorrowRequest.STATUS_RETURNED, 0)

    # Use single aggregated query for user stats
    user_stats = (
        UserProfile.objects
        .filter(role__in=[UserProfile.ROLE_STAFF, UserProfile.ROLE_BORROWER])
        .values('role', 'status')
        .annotate(count=Count('pk'))
    )
    user_stats_dict = {(item['role'], item['status']): item['count'] for item in user_stats}
    
    pending_staff = user_stats_dict.get((UserProfile.ROLE_STAFF, UserProfile.STATUS_PENDING), 0)
    total_staff = user_stats_dict.get((UserProfile.ROLE_STAFF, UserProfile.STATUS_ACTIVE), 0)
    total_borrowers = user_stats_dict.get((UserProfile.ROLE_BORROWER, UserProfile.STATUS_ACTIVE), 0)

    # Optimized queries with select_related and prefetch_related
    recent_borrows = (
        BorrowRequest.objects
        .select_related("user", "approved_by")
        .prefetch_related("items__equipment")
        .order_by("-request_date")[:5]
    )

    recent_returns = (
        ReturnRecord.objects
        .select_related("request__user", "staff")
        .order_by("-borrowed_date")[:5]
    )

    most_borrowed = (
        BorrowRequestItem.objects
        .values("equipment__name")
        .annotate(borrow_count=Count("request_item_id"))
        .order_by("-borrow_count")[:5]
    )

    pending_staff_list = UserProfile.objects.none()
    if profile.role == UserProfile.ROLE_ADMIN:
        pending_staff_list = (
            UserProfile.objects
            .select_related("user")
            .filter(role=UserProfile.ROLE_STAFF, status=UserProfile.STATUS_PENDING)
            .order_by("user__date_joined")
        )

    borrow_month_labels, borrow_month_data = _monthly_counts(
        BorrowRequest.objects.all(), "request_date"
    )
    return_month_labels, return_month_data = _monthly_return_counts()

    recent_activities = (
        AuditLog.objects
        .select_related("user")
        .order_by("-timestamp")[:10]
    )

    context = {
        "profile": profile,
        "total_equipment": total_equipment,
        "available_equipment": available_equipment,
        "unavailable_equipment": unavailable_equipment,
        "maintenance_equipment": maintenance_equipment,
        "low_stock_equipment": low_stock_equipment,
        "out_of_stock_equipment": out_of_stock_equipment,
        "pending_borrows": pending_borrows,
        "approved_borrows": approved_borrows,
        "rejected_borrows": rejected_borrows,
        "returned_borrows": returned_borrows,
        "pending_staff": pending_staff,
        "total_staff": total_staff,
        "total_borrowers": total_borrowers,
        "recent_borrows": recent_borrows,
        "recent_returns": recent_returns,
        "most_borrowed": most_borrowed,
        "pending_staff_list": pending_staff_list,
        "recent_activities": recent_activities,
        "chart_borrow_labels": json.dumps(
            ["Pending", "Approved", "Rejected", "Returned"]
        ),
        "chart_borrow_data": json.dumps(
            [pending_borrows, approved_borrows, rejected_borrows, returned_borrows]
        ),
        "chart_equipment_labels": json.dumps(
            ["Available", "Unavailable", "Maintenance"]
        ),
        "chart_equipment_data": json.dumps(
            [available_equipment, unavailable_equipment, maintenance_equipment]
        ),
        "chart_monthly_borrow_labels": json.dumps(borrow_month_labels),
        "chart_monthly_borrow_data": json.dumps(borrow_month_data),
        "chart_monthly_return_labels": json.dumps(return_month_labels),
        "chart_monthly_return_data": json.dumps(return_month_data),
    }

    return render(request, "dashboard/index.html", context)


# ---------------------------------------------------------------------------
# Equipment CRUD
# ---------------------------------------------------------------------------

@login_required
def equipment_list(request):
    profile = request.user.profile

    search = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()

    equipments = Equipment.objects.all().order_by("name")

    if search:
        equipments = equipments.filter(
            Q(name__icontains=search) |
            Q(category__icontains=search) |
            Q(description__icontains=search)
        )

    if status_filter:
        equipments = equipments.filter(status=status_filter)

    page_obj = _paginate(request, equipments)

    context = {
        "page_obj": page_obj,
        "equipments": page_obj,
        "profile": profile,
        "search": search,
        "status_filter": status_filter,
        "status_choices": Equipment.STATUS_CHOICES,
    }

    return render(request, "equipment/equipment_list.html", context)


@login_required
def equipment_create(request):

    if not _is_admin_or_staff(request.user):
        messages.error(request, "You do not have permission to add equipment.")
        return redirect("equipment_list")

    if request.method == "POST":
        form = EquipmentForm(request.POST, request.FILES)

        if form.is_valid():
            equipment = form.save()
            log_action(
                user=request.user,
                action="equipment_created",
                description=f"Created equipment: {equipment.name} (ID: {equipment.equipment_id})",
                ip_address=request.META.get('REMOTE_ADDR'),
                related_model="Equipment",
                related_id=equipment.equipment_id,
            )
            messages.success(request, "Equipment added successfully.")
            return redirect("equipment_list")
    else:
        form = EquipmentForm()

    return render(
        request,
        "equipment/equipment_form.html",
        {"form": form, "profile": request.user.profile},
    )


@login_required
def equipment_update(request, equipment_id):

    if not _is_admin_or_staff(request.user):
        messages.error(request, "You do not have permission to edit equipment.")
        return redirect("equipment_list")

    equipment = get_object_or_404(Equipment, equipment_id=equipment_id)

    if request.method == "POST":
        form = EquipmentForm(request.POST, request.FILES, instance=equipment)

        if form.is_valid():
            old_name = equipment.name
            form.save()
            log_action(
                user=request.user,
                action="equipment_updated",
                description=f"Updated equipment: {old_name} (ID: {equipment.equipment_id})",
                ip_address=request.META.get('REMOTE_ADDR'),
                related_model="Equipment",
                related_id=equipment.equipment_id,
            )
            messages.success(request, "Equipment updated successfully.")
            return redirect("equipment_detail", equipment_id=equipment.equipment_id)
    else:
        form = EquipmentForm(instance=equipment)

    return render(
        request,
        "equipment/equipment_form.html",
        {
            "form": form,
            "equipment": equipment,
            "profile": request.user.profile,
        },
    )


@login_required
def equipment_detail(request, equipment_id):
    equipment = get_object_or_404(Equipment, equipment_id=equipment_id)

    return render(
        request,
        "equipment/equipment_detail.html",
        {
            "equipment": equipment,
            "profile": request.user.profile,
        },
    )


@login_required
def equipment_delete(request, equipment_id):

    if not _is_admin_or_staff(request.user):
        messages.error(request, "You do not have permission to delete equipment.")
        return redirect("equipment_list")

    equipment = get_object_or_404(Equipment, equipment_id=equipment_id)

    if request.method == "POST":
        equipment_name = equipment.name
        equipment_id = equipment.equipment_id
        equipment.delete()
        log_action(
            user=request.user,
            action="equipment_deleted",
            description=f"Deleted equipment: {equipment_name} (ID: {equipment_id})",
            ip_address=request.META.get('REMOTE_ADDR'),
            related_model="Equipment",
            related_id=equipment_id,
        )
        messages.success(request, "Equipment deleted successfully.")
        return redirect("equipment_list")

    return render(
        request,
        "equipment/equipment_delete.html",
        {
            "equipment": equipment,
            "profile": request.user.profile,
        },
    )


# ---------------------------------------------------------------------------
# Inventory
# ---------------------------------------------------------------------------

@login_required
def inventory_list(request):
    profile = request.user.profile

    search = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()
    stock_filter = request.GET.get("stock", "").strip()
    category_filter = request.GET.get("category", "").strip()
    sort = request.GET.get("sort", "name").strip()

    inventory = Equipment.objects.all()

    if search:
        inventory = inventory.filter(
            Q(name__icontains=search) |
            Q(category__icontains=search) |
            Q(description__icontains=search)
        )

    if status_filter:
        inventory = inventory.filter(status=status_filter)

    if category_filter:
        inventory = inventory.filter(category__iexact=category_filter)

    if stock_filter == "low_stock":
        inventory = inventory.filter(
            quantity_available__gt=0,
            quantity_available__lte=5,
        )
    elif stock_filter == "out_of_stock":
        inventory = inventory.filter(quantity_available=0)

    sort_map = {
        "name": "name",
        "-name": "-name",
        "available": "quantity_available",
        "-available": "-quantity_available",
        "total": "quantity_total",
        "-total": "-quantity_total",
        "category": "category",
        "-category": "-category",
        "status": "status",
        "-status": "-status",
    }
    inventory = inventory.annotate(
        borrowed_quantity=F('quantity_total') - F('quantity_available')
    ).order_by(sort_map.get(sort, "name"))

    # Optimized: Use single aggregated query for all counts
    equipment_stats = (
        Equipment.objects
        .values('status')
        .annotate(count=Count('pk'))
    )
    stats_dict = {item['status']: item['count'] for item in equipment_stats}
    
    total_items = Equipment.objects.count()
    available_items = stats_dict.get(Equipment.STATUS_AVAILABLE, 0)
    unavailable_items = stats_dict.get(Equipment.STATUS_UNAVAILABLE, 0)
    maintenance_items = stats_dict.get(Equipment.STATUS_MAINTENANCE, 0)
    
    low_stock_items = Equipment.objects.filter(
        quantity_available__gt=0,
        quantity_available__lte=5,
    ).count()
    out_of_stock_items = Equipment.objects.filter(
        quantity_available=0
    ).count()

    categories = (
        Equipment.objects
        .values_list("category", flat=True)
        .distinct()
        .order_by("category")
    )

    page_obj = _paginate(request, inventory)

    context = {
        "page_obj": page_obj,
        "inventory": page_obj,
        "profile": profile,
        "search": search,
        "status_filter": status_filter,
        "stock_filter": stock_filter,
        "category_filter": category_filter,
        "sort": sort,
        "categories": categories,
        "status_choices": Equipment.STATUS_CHOICES,
        "total_items": total_items,
        "available_items": available_items,
        "unavailable_items": unavailable_items,
        "maintenance_items": maintenance_items,
        "low_stock_items": low_stock_items,
        "out_of_stock_items": out_of_stock_items,
    }

    return render(request, "inventory/inventory_list.html", context)


@login_required
def low_stock(request):
    profile = request.user.profile

    search = request.GET.get("search", "").strip()
    category_filter = request.GET.get("category", "").strip()
    sort = request.GET.get("sort", "name").strip()

    inventory = Equipment.objects.filter(
        quantity_available__lte=5,
        quantity_available__gt=0,
    )

    if search:
        inventory = inventory.filter(
            Q(name__icontains=search) |
            Q(category__icontains=search) |
            Q(description__icontains=search)
        )

    if category_filter:
        inventory = inventory.filter(category__iexact=category_filter)

    sort_map = {
        "name": "name",
        "-name": "-name",
        "available": "quantity_available",
        "-available": "-quantity_available",
        "total": "quantity_total",
        "-total": "-quantity_total",
        "category": "category",
        "-category": "-category",
    }
    inventory = inventory.annotate(
        borrowed_quantity=F('quantity_total') - F('quantity_available')
    ).order_by(sort_map.get(sort, "name"))

    total_items = Equipment.objects.filter(
        quantity_available__lte=5,
        quantity_available__gt=0,
    ).count()

    categories = (
        Equipment.objects
        .filter(
            quantity_available__lte=5,
            quantity_available__gt=0,
        )
        .values_list("category", flat=True)
        .distinct()
        .order_by("category")
    )

    page_obj = _paginate(request, inventory)

    context = {
        "page_obj": page_obj,
        "inventory": page_obj,
        "profile": profile,
        "search": search,
        "category_filter": category_filter,
        "sort": sort,
        "categories": categories,
        "total_items": total_items,
    }

    return render(request, "inventory/low_stock.html", context)


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------

def _report_context(profile):
    """Helper to build report context with optimized queries."""
    today = timezone.localdate()
    start_of_month = today.replace(day=1)

    # Equipment summary
    equipment_summary = list(
        Equipment.objects
        .values("category")
        .annotate(
            total_quantity=Sum("quantity_total"),
            available_quantity=Sum("quantity_available"),
            equipment_count=Count("pk"),
        )
        .order_by("category")
    )

    # Borrow summary - monthly
    borrow_monthly_summary = list(
        BorrowRequest.objects
        .filter(request_date__date__gte=start_of_month)
        .extra({"date": "date(request_date)"})
        .values("date")
        .annotate(
            pending=Count("pk", filter=Q(status=BorrowRequest.STATUS_PENDING)),
            approved=Count("pk", filter=Q(status=BorrowRequest.STATUS_APPROVED)),
            rejected=Count("pk", filter=Q(status=BorrowRequest.STATUS_REJECTED)),
            returned=Count("pk", filter=Q(status=BorrowRequest.STATUS_RETURNED)),
        )
        .order_by("date")
    )

    # Most borrowed equipment
    most_borrowed = list(
        BorrowRequestItem.objects
        .values("equipment__name", "equipment__category")
        .annotate(borrow_count=Count("request_item_id"))
        .order_by("-borrow_count")[:10]
    )

    # Most active borrowers
    most_active_borrowers = list(
        BorrowRequest.objects
        .values("user__username", "user__first_name", "user__last_name")
        .annotate(request_count=Count("request_id"))
        .order_by("-request_count")[:10]
    )

    # Category statistics
    category_stats = list(
        BorrowRequestItem.objects
        .values("equipment__category")
        .annotate(
            borrow_count=Count("request_item_id"),
            total_quantity=Sum("quantity"),
        )
        .order_by("-borrow_count")
    )

    return {
        "equipment_summary": equipment_summary,
        "borrow_monthly_summary": borrow_monthly_summary,
        "most_borrowed": most_borrowed,
        "most_active_borrowers": most_active_borrowers,
        "category_stats": category_stats,
        "start_of_month": start_of_month,
        "today": today,
    }


@login_required
def reports(request):
    profile = request.user.profile

    context = _report_context(profile)
    context["profile"] = profile

    return render(request, "reports/reports.html", context)


@login_required
def reports_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from django.http import HttpResponse

    profile = request.user.profile
    context = _report_context(profile)

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Report"

    # Write headers
    ws.append(["Equipment Report", "", ""])
    ws.append(["Generated on:", timezone.localdate().strftime("%Y-%m-%d"), ""])
    ws.append([])
    ws.append(["Equipment Summary", "", ""])
    ws.append(["Category", "Total Quantity", "Available", "Borrowed", "Count"])

    for row in context["equipment_summary"]:
        borrowed = row["total_quantity"] - row["available_quantity"]
        ws.append([
            row["category"] or "Uncategorized",
            row["total_quantity"] or 0,
            row["available_quantity"] or 0,
            borrowed,
            row["equipment_count"],
        ])

    ws.append([])
    ws.append(["Most Borrowed Equipment", "", ""])
    ws.append(["Equipment", "Category", "Borrow Count"])
    for row in context["most_borrowed"]:
        ws.append([row["equipment__name"], row["equipment__category"], row["borrow_count"]])

    ws.append([])
    ws.append(["Most Active Borrowers", "", ""])
    ws.append(["Username", "Name", "Request Count"])
    for row in context["most_active_borrowers"]:
        name = f"{row['user__first_name'] or ''} {row['user__last_name'] or ''}".strip()
        ws.append([row["user__username"], name, row["request_count"]])

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for cell in ws[5]:
        cell.font = Font(bold=True)
    for cell in ws[10]:
        cell.font = Font(bold=True)
    for cell in ws[14]:
        cell.font = Font(bold=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=equipment_report.xlsx"
    wb.save(response)

    log_action(
        user=request.user,
        action="report_exported",
        description="Exported Excel report",
        ip_address=request.META.get('REMOTE_ADDR'),
        related_model="Report",
        related_id=0,
    )

    return response


@login_required
def reports_export_pdf(request):
    from reportlab.lib.pagesizes import letter, inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO

    profile = request.user.profile
    context = _report_context(profile)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=equipment_report.pdf"

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=24,
        textColor=colors.HexColor("#2c3e50"),
        spaceAfter=30,
        alignment=TA_CENTER,
    )
    story.append(Paragraph("Equipment Report", title_style))
    story.append(Spacer(1, 0.1 * inch))
    story.append(
        Paragraph(
            f"Generated on: {timezone.localdate().strftime('%Y-%m-%d')}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.3 * inch))

    # Equipment Summary Table
    story.append(Paragraph("Equipment Summary", styles["Heading2"]))
    story.append(Spacer(1, 0.1 * inch))

    equipment_data = [
        ["Category", "Total", "Available", "Borrowed", "Count"]
    ]
    for row in context["equipment_summary"]:
        borrowed = (row["total_quantity"] or 0) - (row["available_quantity"] or 0)
        equipment_data.append([
            row["category"] or "Uncategorized",
            row["total_quantity"] or 0,
            row["available_quantity"] or 0,
            borrowed,
            row["equipment_count"],
        ])

    equipment_table = Table(equipment_data)
    equipment_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ])
    )
    story.append(equipment_table)
    story.append(Spacer(1, 0.3 * inch))

    # Most Borrowed Equipment
    story.append(Paragraph("Most Borrowed Equipment", styles["Heading2"]))
    story.append(Spacer(1, 0.1 * inch))

    borrowed_data = [["Equipment", "Category", "Borrow Count"]]
    for row in context["most_borrowed"]:
        borrowed_data.append([
            row["equipment__name"],
            row["equipment__category"],
            row["borrow_count"],
        ])

    borrowed_table = Table(borrowed_data)
    borrowed_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ])
    )
    story.append(borrowed_table)
    story.append(Spacer(1, 0.3 * inch))

    # Most Active Borrowers
    story.append(Paragraph("Most Active Borrowers", styles["Heading2"]))
    story.append(Spacer(1, 0.1 * inch))

    borrowers_data = [["Username", "Name", "Request Count"]]
    for row in context["most_active_borrowers"]:
        name = f"{row['user__first_name'] or ''} {row['user__last_name'] or ''}".strip()
        borrowers_data.append([row["user__username"], name, row["request_count"]])

    borrowers_table = Table(borrowers_data)
    borrowers_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ])
    )
    story.append(borrowers_table)

    doc.build(story)
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()

    log_action(
        user=request.user,
        action="report_exported",
        description="Exported PDF report",
        ip_address=request.META.get('REMOTE_ADDR'),
        related_model="Report",
        related_id=0,
    )

    return response


# ---------------------------------------------------------------------------
# Borrow Requests
# ---------------------------------------------------------------------------

@login_required
def borrow_list(request):
    profile = request.user.profile

    search = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()

    if profile.role == UserProfile.ROLE_BORROWER:
        borrow_requests = BorrowRequest.objects.filter(
            user=request.user
        ).select_related("user", "approved_by").prefetch_related("items__equipment")
    else:
        borrow_requests = BorrowRequest.objects.all().select_related(
            "user", "approved_by"
        ).prefetch_related("items__equipment")

    if search:
        search_q = (
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__username__icontains=search)
        )
        if search.isdigit():
            search_q |= Q(request_id=int(search))
        borrow_requests = borrow_requests.filter(search_q)

    if status_filter:
        borrow_requests = borrow_requests.filter(status=status_filter)

    page_obj = _paginate(request, borrow_requests)

    return render(
        request,
        "borrow/borrow_list.html",
        {
            "page_obj": page_obj,
            "borrow_requests": page_obj,
            "profile": profile,
            "search": search,
            "status_filter": status_filter,
            "status_choices": BorrowRequest.STATUS_CHOICES,
        },
    )


@login_required
def borrow_create(request):
    profile = request.user.profile

    if profile.role != UserProfile.ROLE_BORROWER:
        messages.error(request, "Only borrowers can submit borrow requests.")
        return redirect("borrow_list")

    if request.method == "POST":
        formset = BorrowRequestItemFormSet(request.POST)

        if formset.is_valid():
            with transaction.atomic():
                # Create the borrow request
                borrow_request = BorrowRequest.objects.create(
                    user=request.user,
                    status=BorrowRequest.STATUS_PENDING,
                )

                # Create request items
                for form in formset:
                    if form.cleaned_data:
                        item = form.save(commit=False)
                        item.request = borrow_request
                        item.save()

                log_action(
                    user=request.user,
                    action="borrow_created",
                    description=f"Created borrow request #{borrow_request.request_id}",
                    ip_address=request.META.get('REMOTE_ADDR'),
                    related_model="BorrowRequest",
                    related_id=borrow_request.request_id,
                )
                messages.success(
                    request,
                    f"Borrow request #{borrow_request.request_id} created successfully. "
                    "Waiting for staff approval.",
                )
                return redirect("borrow_list")
    else:
        formset = BorrowRequestItemFormSet()

    # Get available equipment for the form
    available_equipment = Equipment.objects.filter(
        status=Equipment.STATUS_AVAILABLE,
        quantity_available__gt=0,
    ).order_by("name")

    return render(
        request,
        "borrow/borrow_form.html",
        {
            "formset": formset,
            "profile": profile,
            "available_equipment": available_equipment,
        },
    )


@login_required
def borrow_detail(request, request_id):
    profile = request.user.profile

    if profile.role == UserProfile.ROLE_BORROWER:
        borrow_request = get_object_or_404(
            BorrowRequest.objects.select_related("user", "approved_by").prefetch_related("items__equipment"),
            request_id=request_id,
            user=request.user,
        )
    else:
        borrow_request = get_object_or_404(
            BorrowRequest.objects.select_related("user", "approved_by").prefetch_related("items__equipment"),
            request_id=request_id,
        )

    return render(
        request,
        "borrow/borrow_detail.html",
        {
            "borrow_request": borrow_request,
            "items": borrow_request.items.all(),
            "profile": profile,
        },
    )


@login_required
def borrow_delete(request, request_id):
    profile = request.user.profile

    if profile.role != UserProfile.ROLE_BORROWER:
        messages.error(request, "Only borrowers can delete their own requests.")
        return redirect("borrow_list")

    borrow_request = get_object_or_404(
        BorrowRequest.objects.select_related("user"),
        request_id=request_id,
        user=request.user,
        status=BorrowRequest.STATUS_PENDING,
    )

    if request.method == "POST":
        request_id = borrow_request.request_id
        borrow_request.delete()

        log_action(
            user=request.user,
            action="borrow_cancelled",
            description=f"Cancelled borrow request #{request_id}",
            ip_address=request.META.get('REMOTE_ADDR'),
            related_model="BorrowRequest",
            related_id=request_id,
        )
        messages.success(request, f"Borrow request #{request_id} cancelled.")
        return redirect("borrow_list")

    return render(
        request,
        "borrow/borrow_confirm_delete.html",
        {
            "borrow_request": borrow_request,
            "profile": profile,
        },
    )


@login_required
def approve_borrow(request, request_id):
    if not _is_admin(request.user):
        messages.error(request, "Only admins can approve borrow requests.")
        return redirect("borrow_list")

    borrow_request = get_object_or_404(
        BorrowRequest.objects.select_related("user").prefetch_related("items__equipment"),
        request_id=request_id,
        status=BorrowRequest.STATUS_PENDING,
    )

    if request.method == "POST":
        with transaction.atomic():
            for item in borrow_request.items.all():
                equipment = item.equipment
                if item.quantity > equipment.quantity_available:
                    messages.error(
                        request,
                        f'Not enough stock for "{equipment.name}". '
                        f"Available: {equipment.quantity_available}, "
                        f"Requested: {item.quantity}.",
                    )
                    return redirect("borrow_detail", request_id=request_id)

                equipment.quantity_available -= item.quantity
                equipment.save()

            borrow_request.status = BorrowRequest.STATUS_APPROVED
            borrow_request.approved_by = request.user
            borrow_request.approved_at = timezone.now()
            borrow_request.save()

        log_action(
            user=request.user,
            action="borrow_approved",
            description=f"Approved borrow request #{request_id} for {borrow_request.user.username}",
            ip_address=request.META.get('REMOTE_ADDR'),
            related_model="BorrowRequest",
            related_id=request_id,
        )
        messages.success(
            request,
            f"Borrow request #{request_id} approved successfully.",
        )

    return redirect("borrow_list")


@login_required
def reject_borrow(request, request_id):
    if not _is_admin(request.user):
        messages.error(request, "Only admins can reject borrow requests.")
        return redirect("borrow_list")

    borrow_request = get_object_or_404(
        BorrowRequest.objects.select_related("user"),
        request_id=request_id,
        status=BorrowRequest.STATUS_PENDING,
    )

    if request.method == "POST":
        borrow_request.status = BorrowRequest.STATUS_REJECTED
        borrow_request.approved_by = request.user
        borrow_request.approved_at = timezone.now()
        borrow_request.save()

        log_action(
            user=request.user,
            action="borrow_rejected",
            description=f"Rejected borrow request #{request_id} for {borrow_request.user.username}",
            ip_address=request.META.get('REMOTE_ADDR'),
            related_model="BorrowRequest",
            related_id=request_id,
        )
        messages.warning(request, f"Borrow request #{request_id} rejected.")

    return redirect("borrow_list")


# ---------------------------------------------------------------------------
# Returns
# ---------------------------------------------------------------------------

@login_required
def return_list(request):
    profile = request.user.profile

    search = request.GET.get("search", "").strip()
    filter_type = request.GET.get("filter", "").strip()

    if profile.role == UserProfile.ROLE_BORROWER:
        returns = ReturnRecord.objects.filter(
            request__user=request.user
        ).select_related("request__user", "staff")
    else:
        returns = ReturnRecord.objects.all().select_related(
            "request__user", "staff"
        )

    if search:
        search_q = (
            Q(request__user__first_name__icontains=search) |
            Q(request__user__last_name__icontains=search) |
            Q(request__user__username__icontains=search)
        )
        if search.isdigit():
            search_q |= Q(transaction_id=int(search))
        returns = returns.filter(search_q)

    today = timezone.localdate()

    if filter_type == "overdue":
        returns = returns.filter(
            return_date__isnull=True,
            due_date__lt=today,
        )
    elif filter_type == "completed":
        returns = returns.filter(return_date__isnull=False)

    page_obj = _paginate(request, returns)

    return render(
        request,
        "return/return_list.html",
        {
            "page_obj": page_obj,
            "returns": page_obj,
            "profile": profile,
            "search": search,
            "filter_type": filter_type,
            "today": today,
        },
    )


@login_required
def create_return(request, request_id):
    if not _is_admin_or_staff(request.user):
        messages.error(request, "Only staff or admins can process returns.")
        return redirect("borrow_list")

    borrow_request = get_object_or_404(
        BorrowRequest.objects.select_related("user").prefetch_related("items__equipment"),
        request_id=request_id,
        status=BorrowRequest.STATUS_APPROVED,
    )

    if hasattr(borrow_request, "return_record"):
        messages.info(request, "This request has already been returned.")
        return redirect("return_list")

    if request.method == "POST":
        form = ReturnRecordForm(request.POST)

        if form.is_valid():
            with transaction.atomic():
                return_record = form.save(commit=False)
                return_record.request = borrow_request
                return_record.staff = request.user
                return_record.save()

                for item in borrow_request.items.all():
                    equipment = item.equipment
                    equipment.quantity_available += item.quantity
                    equipment.save()

                borrow_request.status = BorrowRequest.STATUS_RETURNED
                borrow_request.save()

            log_action(
                user=request.user,
                action="borrow_returned",
                description=f"Processed return for borrow request #{request_id} by {borrow_request.user.username}",
                ip_address=request.META.get('REMOTE_ADDR'),
                related_model="BorrowRequest",
                related_id=request_id,
            )
            messages.success(
                request,
                f"Return for request #{request_id} recorded successfully.",
            )
            return redirect("return_list")

    else:
        form = ReturnRecordForm(
            initial={"borrowed_date": timezone.localdate()}
        )

    return render(
        request,
        "return/return_form.html",
        {
            "form": form,
            "borrow_request": borrow_request,
            "items": borrow_request.items.all(),
            "profile": request.user.profile,
        },
    )


# ---------------------------------------------------------------------------
# User Profile
# ---------------------------------------------------------------------------

@login_required
def profile_view(request):
    profile = request.user.profile
    borrow_requests = (
        BorrowRequest.objects.filter(user=request.user)
        .select_related("user", "approved_by")
        .prefetch_related("items__equipment")
        .order_by("-request_date")[:10]
    )

    return render(
        request,
        "accounts/profile.html",
        {
            "profile": profile,
            "borrow_requests": borrow_requests,
        },
    )


@login_required
def edit_profile(request):
    profile = request.user.profile

    if request.method == "POST":
        form = EditProfileForm(request.POST, instance=profile)

        if form.is_valid():
            old_department = profile.department
            form.save()

            log_action(
                user=request.user,
                action="profile_updated",
                description=f"Updated profile: department '{old_department}' to '{profile.department}'",
                ip_address=request.META.get('REMOTE_ADDR'),
                related_model="UserProfile",
                related_id=profile.user_id,
            )
            messages.success(request, "Profile updated successfully.")
            return redirect("profile")
    else:
        form = EditProfileForm(instance=profile)

    return render(
        request,
        "accounts/edit_profile.html",
        {"form": form, "profile": profile},
    )


@login_required
def change_password(request):
    if request.method == "POST":
        form = PasswordChangeForm(request.user, request.POST)

        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)

            log_action(
                user=request.user,
                action="password_changed",
                description="User changed password",
                ip_address=request.META.get('REMOTE_ADDR'),
                related_model="User",
                related_id=request.user.id,
            )
            messages.success(request, "Password changed successfully.")
            return redirect("profile")
    else:
        form = PasswordChangeForm(request.user)

    return render(
        request,
        "accounts/change_password.html",
        {"form": form, "profile": request.user.profile},
    )


@login_required
def user_list(request):
    if not _is_admin(request.user):
        messages.error(request, "Only administrators can manage users.")
        return redirect("dashboard")

    search = request.GET.get("search", "").strip()
    role_filter = request.GET.get("role", "").strip()
    status_filter = request.GET.get("status", "").strip()

    users = UserProfile.objects.select_related("user").order_by(
        "-user__date_joined"
    )

    if search:
        users = users.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(contact_number__icontains=search) |
            Q(department__icontains=search)
        )

    if role_filter:
        users = users.filter(role=role_filter)

    if status_filter:
        users = users.filter(status=status_filter)

    page_obj = _paginate(request, users)

    return render(
        request,
        "accounts/user_list.html",
        {
            "page_obj": page_obj,
            "users": page_obj,
            "profile": request.user.profile,
            "search": search,
            "role_filter": role_filter,
            "status_filter": status_filter,
            "role_choices": UserProfile.ROLE_CHOICES,
            "status_choices": UserProfile.STATUS_CHOICES,
        },
    )
